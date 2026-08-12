# -*- coding: utf-8 -*-
"""
سامانهٔ حضور و غیاب با تشخیص چهره — بک‌اند Flask.

تمام پردازش تصویر و تشخیص چهره در face_engine.py (با OpenCV) و سمت سرور
انجام می‌شود. مرورگر فقط فریم وب‌کم را به‌صورت تصویر می‌فرستد.
"""

import io
import random
from datetime import datetime, timedelta

from flask import (
    Flask, render_template, request, jsonify, send_file, abort,
)

import storage
import reports as reports_mod
from face_engine import FaceEngine, FaceEngineError

app = Flask(__name__)

# ---------- مقداردهی موتور تشخیص چهره ----------
# موتور را همین ابتدا (هنگام بالا آمدن سرور) می‌سازیم تا مدل‌ها یک‌بار در حافظه بارگذاری
# شوند و هر درخواست سریع باشد. اگر مدل‌ها نباشند یا OpenCV نصب نباشد، برنامه را crash
# نمی‌کنیم؛ فقط پیام خطا را نگه می‌داریم و بقیهٔ بخش‌ها (کارمندان، گزارش‌ها...) کار می‌کنند.
ENGINE = None
ENGINE_ERROR = None
try:
    ENGINE = FaceEngine(cosine_threshold=storage.get_settings().get("threshold", 0.363))
except FaceEngineError as e:
    ENGINE_ERROR = str(e)
except Exception as e:  # کتابخانهٔ OpenCV نصب نیست یا مشکل دیگر
    ENGINE_ERROR = "موتور تشخیص چهره راه‌اندازی نشد: " + str(e)


def require_engine():
    # نگهبان مسیرهایی که به موتور نیاز دارند (ثبت چهره و تشخیص). اگر موتور آماده نباشد
    # پاسخ خطا برمی‌گردانیم. ضمناً آستانه را از تنظیمات تازه می‌خوانیم تا اگر کاربر همین
    # حالا در صفحهٔ تنظیمات آن را تغییر داده باشد، بدون ری‌استارت سرور اعمال شود.
    if ENGINE is None:
        return jsonify(ok=False, error=ENGINE_ERROR or "موتور تشخیص چهره در دسترس نیست."), 503
    ENGINE.cosine_threshold = float(storage.get_settings().get("threshold", ENGINE.cosine_threshold))
    return None


# =====================================================================
#  صفحه‌ها
# =====================================================================
PAGES = ["dashboard", "employees", "register", "kiosk", "reports", "settings"]


@app.route("/")
def page_dashboard():
    return render_template("dashboard.html", active="dashboard", engine_error=ENGINE_ERROR)


@app.route("/employees")
def page_employees():
    return render_template("employees.html", active="employees", engine_error=ENGINE_ERROR)


@app.route("/register")
def page_register():
    return render_template("register.html", active="register", engine_error=ENGINE_ERROR)


@app.route("/kiosk")
def page_kiosk():
    return render_template("kiosk.html", active="kiosk", engine_error=ENGINE_ERROR)


@app.route("/reports")
def page_reports():
    return render_template("reports.html", active="reports", engine_error=ENGINE_ERROR)


@app.route("/settings")
def page_settings():
    return render_template("settings.html", active="settings", engine_error=ENGINE_ERROR)


# =====================================================================
#  کمک‌تابع‌ها
# =====================================================================
def public_employee(e):
    """نسخهٔ سبکِ کارمند برای ارسال به کلاینت (بدون بردارهای سنگین چهره).

    عمداً فیلد encodings را به مرورگر نمی‌فرستیم: هم حجم پاسخ کم می‌شود و هم دادهٔ
    بیومتریک روی سرور می‌ماند. فقط تعداد نمونه‌ها (samples) را اعلام می‌کنیم.
    """
    return {
        "id": e["id"],
        "code": e.get("code", ""),
        "name": e.get("name", ""),
        "position": e.get("position", ""),
        "unit": e.get("unit", ""),
        "status": e.get("status", "فعال"),
        "samples": len(e.get("encodings", [])),
        "has_face": len(e.get("encodings", [])) > 0,
    }


def today_jdate():
    return storage.jalali_date_str(storage.now_ms())


def employee_today_state(att, emp_id):
    """آخرین وضعیت امروزِ کارمند: in / out / none + زمان آخرین رویداد.

    منطق ساده است: آخرین رویداد امروز را پیدا می‌کنیم. اگر «ورود» بود یعنی فرد داخل است
    و دفعهٔ بعد باید «خروج» بزند، و برعکس. کیوسک از همین برای پیشنهاد دکمهٔ درست استفاده می‌کند.
    """
    tj = today_jdate()
    rows = [a for a in att if a["employee_id"] == emp_id and a.get("jdate") == tj]
    if not rows:
        return "none", None
    rows.sort(key=lambda a: a["ts"])   # بر اساس زمان مرتب کن تا «آخرین» رویداد درست باشد
    last = rows[-1]
    return last["type"], storage.time_str(last["ts"])


# =====================================================================
#  API — وضعیت کلی
# =====================================================================
@app.route("/api/state")
def api_state():
    s = storage.get_settings()
    emps = storage.load("employees")
    return jsonify(
        ok=True,
        engine_ready=ENGINE is not None,
        engine_error=ENGINE_ERROR,
        settings=s,
        employee_count=len(emps),
        registered_faces=sum(1 for e in emps if e.get("encodings")),
    )


@app.route("/api/dashboard")
def api_dashboard():
    emps = storage.load("employees")
    att = storage.load("attendance")
    tj = today_jdate()
    today_rows = [a for a in att if a.get("jdate") == tj]

    present_ids = {a["employee_id"] for a in today_rows if a["type"] == "in"}
    s = storage.get_settings()
    start_min = reports_mod._to_minutes(s.get("start_time", "08:30"))

    # «تأخیر» را بر اساس اولین ورودِ هر شخص می‌سنجیم: اگر اولین ورود امروز بعد از ساعت
    # شروع کار باشد، آن فرد تأخیر داشته. (ورودهای بعدیِ همان روز را برای تأخیر نمی‌شماریم.)
    late_ids = set()
    by_emp = {}
    for a in sorted(today_rows, key=lambda x: x["ts"]):
        by_emp.setdefault(a["employee_id"], []).append(a)
    for emp_id, rows in by_emp.items():
        ins = [r for r in rows if r["type"] == "in"]
        if ins:
            # ts بر حسب میلی‌ثانیه است؛ این فرمول دقیقهٔ روز (۰ تا ۱۴۳۹) را می‌دهد.
            m = (ins[0]["ts"] // 60000) % (24 * 60)
            if m > start_min:
                late_ids.add(emp_id)

    emap = {e["id"]: e for e in emps}
    recent = sorted(today_rows, key=lambda x: x["ts"], reverse=True)[:12]
    recent_out = [{
        "name": emap.get(r["employee_id"], {}).get("name", "—"),
        "code": emap.get(r["employee_id"], {}).get("code", ""),
        "type": r["type"],
        "time": storage.time_str(r["ts"]),
    } for r in recent]

    total = len(emps)
    return jsonify(
        ok=True,
        total_employees=total,
        present_today=len(present_ids),
        late_today=len(late_ids),
        absent_today=max(0, total - len(present_ids)),
        recent=recent_out,
    )


# =====================================================================
#  API — کارمندان
# =====================================================================
@app.route("/api/employees", methods=["GET"])
def api_employees_list():
    emps = storage.load("employees")
    return jsonify(ok=True, employees=[public_employee(e) for e in emps])


@app.route("/api/employees", methods=["POST"])
def api_employees_create():
    d = request.get_json(force=True) or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify(ok=False, error="نام کارمند الزامی است."), 400
    emps = storage.load("employees")
    code = (d.get("code") or "").strip()
    if code and any(e.get("code") == code for e in emps):
        return jsonify(ok=False, error="این کد پرسنلی قبلاً ثبت شده است."), 400
    emp = {
        "id": storage.new_id(),
        "code": code,
        "name": name,
        "position": (d.get("position") or "").strip(),
        "unit": (d.get("unit") or "").strip(),
        "status": (d.get("status") or "فعال").strip(),
        "encodings": [],
        "created": storage.now_ms(),
    }
    emps.append(emp)
    storage.save("employees", emps)
    return jsonify(ok=True, employee=public_employee(emp))


@app.route("/api/employees/<emp_id>", methods=["PUT"])
def api_employees_update(emp_id):
    d = request.get_json(force=True) or {}
    emps = storage.load("employees")
    for e in emps:
        if e["id"] == emp_id:
            for k in ("code", "name", "position", "unit", "status"):
                if k in d and d[k] is not None:
                    e[k] = str(d[k]).strip()
            storage.save("employees", emps)
            return jsonify(ok=True, employee=public_employee(e))
    return jsonify(ok=False, error="کارمند یافت نشد."), 404


@app.route("/api/employees/<emp_id>", methods=["DELETE"])
def api_employees_delete(emp_id):
    emps = storage.load("employees")
    new = [e for e in emps if e["id"] != emp_id]
    if len(new) == len(emps):
        return jsonify(ok=False, error="کارمند یافت نشد."), 404
    storage.save("employees", new)
    # وقتی کارمندی حذف می‌شود، ترددهای ثبت‌شدهٔ او را هم پاک می‌کنیم تا رکوردهای بی‌صاحب
    # در گزارش‌ها باقی نماند و آمار به‌هم نریزد.
    att = storage.load("attendance")
    storage.save("attendance", [a for a in att if a["employee_id"] != emp_id])
    return jsonify(ok=True)


# =====================================================================
#  API — ثبت چهره
# =====================================================================
@app.route("/api/face/register", methods=["POST"])
def api_face_register():
    guard = require_engine()
    if guard:
        return guard
    d = request.get_json(force=True) or {}
    emp_id = d.get("employee_id")
    image = d.get("image")
    emps = storage.load("employees")
    emp = next((e for e in emps if e["id"] == emp_id), None)
    if not emp:
        return jsonify(ok=False, error="ابتدا کارمند را انتخاب کنید."), 400
    try:
        feature, box = ENGINE.encode_single(image)
    except FaceEngineError as e:
        return jsonify(ok=False, error=str(e)), 422
    emp.setdefault("encodings", []).append(feature)
    # فقط ۵ نمونهٔ آخر را نگه می‌داریم: چند نمونه از زاویه‌های مختلف دقت را بالا می‌برد،
    # ولی نگه‌داشتن تعداد زیاد، هم مقایسه را کند می‌کند و هم سود چندانی ندارد.
    emp["encodings"] = emp["encodings"][-5:]
    storage.save("employees", emps)
    return jsonify(ok=True, samples=len(emp["encodings"]), box=box)


@app.route("/api/face/<emp_id>", methods=["DELETE"])
def api_face_clear(emp_id):
    emps = storage.load("employees")
    emp = next((e for e in emps if e["id"] == emp_id), None)
    if not emp:
        return jsonify(ok=False, error="کارمند یافت نشد."), 404
    emp["encodings"] = []
    storage.save("employees", emps)
    return jsonify(ok=True, samples=0)


# =====================================================================
#  API — تشخیص زنده (کیوسک)
# =====================================================================
@app.route("/api/recognize", methods=["POST"])
def api_recognize():
    guard = require_engine()
    if guard:
        return guard
    d = request.get_json(force=True) or {}
    image = d.get("image")
    emps = storage.load("employees")
    # فهرست «چهره‌های شناخته‌شده» را صاف می‌کنیم: هر نمونهٔ هر کارمند یک ردیف (employee_id, بردار)
    # می‌شود، چون یک کارمند ممکن است چند نمونهٔ چهره داشته باشد.
    known = [(e["id"], enc) for e in emps for enc in e.get("encodings", [])]
    try:
        result = ENGINE.recognize(image, known)
    except FaceEngineError as e:
        return jsonify(ok=False, error=str(e)), 422

    match = None
    if result["employee_id"]:
        emp = next((e for e in emps if e["id"] == result["employee_id"]), None)
        if emp:
            att = storage.load("attendance")
            state, last_time = employee_today_state(att, emp["id"])
            # دکمهٔ پیشنهادی برعکسِ وضعیت فعلی است: اگر داخل است → خروج، وگرنه → ورود.
            suggested = "out" if state == "in" else "in"
            match = {
                "id": emp["id"],
                "code": emp.get("code", ""),
                "name": emp.get("name", ""),
                "position": emp.get("position", ""),
                "unit": emp.get("unit", ""),
                "today_state": state,
                "suggested": suggested,
                "last_time": last_time,
            }
    return jsonify(
        ok=True,
        face_found=result["face_found"],
        box=result["box"],
        score=result["score"],
        match=match,
    )


# =====================================================================
#  API — ثبت تردد
# =====================================================================
@app.route("/api/attendance", methods=["POST"])
def api_attendance_log():
    d = request.get_json(force=True) or {}
    emp_id = d.get("employee_id")
    typ = d.get("type")
    if typ not in ("in", "out"):
        return jsonify(ok=False, error="نوع تردد نامعتبر است."), 400
    emps = storage.load("employees")
    emp = next((e for e in emps if e["id"] == emp_id), None)
    if not emp:
        return jsonify(ok=False, error="کارمند یافت نشد."), 404

    att = storage.load("attendance")
    now = storage.now_ms()
    # محافظ ثبت تکراری: چون کیوسک هر چند ثانیه یک‌بار چهره را تشخیص می‌دهد، ممکن است کاربر
    # ناخواسته دو بار پشت‌سر‌هم دکمه را بزند. اگر همان کارمند همان نوع تردد را در ۹۰ ثانیهٔ
    # اخیر ثبت کرده باشد، تردد دوم را رد می‌کنیم تا رکورد تکراری ساخته نشود.
    for a in att:
        if a["employee_id"] == emp_id and a["type"] == typ and now - a["ts"] < 90_000:
            return jsonify(ok=False, error="این تردد همین الان ثبت شده است."), 409

    rec = {
        "id": storage.new_id(),
        "employee_id": emp_id,
        "type": typ,
        "ts": now,
        "jdate": storage.jalali_date_str(now),
    }
    att.append(rec)
    storage.save("attendance", att)
    state, last_time = employee_today_state(att, emp_id)
    return jsonify(
        ok=True,
        record={"type": typ, "time": storage.time_str(now)},
        name=emp.get("name", ""),
        today_state=state,
        suggested=("out" if state == "in" else "in"),
    )


# =====================================================================
#  API — گزارش حضور و غیاب
# =====================================================================
@app.route("/api/reports")
def api_reports():
    from_j = request.args.get("from")
    to_j = request.args.get("to")
    if not from_j or not to_j:
        from_j, to_j = reports_mod.current_month_range()
    unit = request.args.get("unit") or None
    emp_id = request.args.get("employee") or None

    emps = storage.load("employees")
    att = storage.load("attendance")
    s = storage.get_settings()
    try:
        rep = reports_mod.build_report(emps, att, s, from_j, to_j, unit, emp_id)
    except Exception as e:
        return jsonify(ok=False, error="بازهٔ تاریخ نامعتبر است: " + str(e)), 400

    # وضعیت تأیید هر سطر را ضمیمه می‌کنیم. کلید تأیید ترکیبی از (کارمند + بازهٔ تاریخ) است،
    # یعنی تأیید یک نفر برای یک ماه، با ماه دیگر قاطی نمی‌شود. اگر چیزی ثبت نشده باشد «pending».
    approvals = storage.load("approvals")
    amap = {(a["employee_id"], a["from"], a["to"]): a["status"] for a in approvals}
    for r in rep["rows"]:
        r["approval"] = amap.get((r["id"], from_j, to_j), "pending")

    units = sorted({e.get("unit", "") for e in emps if e.get("unit")})
    return jsonify(ok=True, report=rep, units=units,
                   employees=[{"id": e["id"], "name": e["name"]} for e in emps])


@app.route("/api/reports/approve", methods=["POST"])
def api_reports_approve():
    d = request.get_json(force=True) or {}
    emp_id, from_j, to_j = d.get("employee_id"), d.get("from"), d.get("to")
    status = d.get("status")
    if status not in ("approved", "rejected", "pending"):
        return jsonify(ok=False, error="وضعیت نامعتبر است."), 400
    approvals = storage.load("approvals")
    approvals = [a for a in approvals
                 if not (a["employee_id"] == emp_id and a["from"] == from_j and a["to"] == to_j)]
    if status != "pending":
        approvals.append({"id": storage.new_id(), "employee_id": emp_id,
                          "from": from_j, "to": to_j, "status": status, "ts": storage.now_ms()})
    storage.save("approvals", approvals)
    return jsonify(ok=True, status=status)


@app.route("/api/reports/export")
def api_reports_export():
    """خروجی اکسل (.xlsx) از گزارش حضور و غیاب — ساخته‌شده در پایتون با openpyxl."""
    from_j = request.args.get("from") or reports_mod.current_month_range()[0]
    to_j = request.args.get("to") or reports_mod.current_month_range()[1]
    unit = request.args.get("unit") or None
    emp_id = request.args.get("employee") or None

    emps = storage.load("employees")
    att = storage.load("attendance")
    s = storage.get_settings()
    rep = reports_mod.build_report(emps, att, s, from_j, to_j, unit, emp_id)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        return jsonify(ok=False, error="برای خروجی اکسل باید کتابخانهٔ openpyxl نصب باشد."), 503

    wb = Workbook()
    ws = wb.active
    ws.title = "حضور و غیاب"
    ws.sheet_view.rightToLeft = True   # برگه را راست‌به‌چپ می‌کنیم تا برای فارسی درست بچیند
    headers = ["کد پرسنلی", "نام", "واحد", "روزهای کاری", "حاضر", "غایب", "تأخیر", "اضافه‌کاری (ساعت)"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="0D1526")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = head_fill
    for r in rep["rows"]:
        ws.append([r["code"], r["name"], r["unit"], r["working_days"],
                   r["present"], r["absent"], r["late"], r["overtime"]])
    widths = [14, 22, 16, 13, 9, 9, 9, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "attendance_%s_%s.xlsx" % (from_j, to_j)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# =====================================================================
#  API — مرخصی‌ها
# =====================================================================
@app.route("/api/leaves", methods=["GET"])
def api_leaves_list():
    leaves = storage.load("leaves")
    emap = {e["id"]: e for e in storage.load("employees")}
    out = []
    for lv in sorted(leaves, key=lambda x: x.get("from", ""), reverse=True):
        e = emap.get(lv["employee_id"], {})
        out.append({**lv, "name": e.get("name", "—"), "code": e.get("code", "")})
    return jsonify(ok=True, leaves=out)


@app.route("/api/leaves", methods=["POST"])
def api_leaves_create():
    d = request.get_json(force=True) or {}
    if not d.get("employee_id") or not d.get("from") or not d.get("to"):
        return jsonify(ok=False, error="کارمند و بازهٔ تاریخ الزامی است."), 400
    leaves = storage.load("leaves")
    lv = {
        "id": storage.new_id(),
        "employee_id": d["employee_id"],
        "from": d["from"].strip(),
        "to": d["to"].strip(),
        "type": (d.get("type") or "استحقاقی").strip(),
        "days": d.get("days") or "",
        "note": (d.get("note") or "").strip(),
        "status": "ثبت‌شده",
    }
    leaves.append(lv)
    storage.save("leaves", leaves)
    return jsonify(ok=True, leave=lv)


@app.route("/api/leaves/<lid>", methods=["DELETE"])
def api_leaves_delete(lid):
    leaves = storage.load("leaves")
    new = [l for l in leaves if l["id"] != lid]
    if len(new) == len(leaves):
        return jsonify(ok=False, error="مرخصی یافت نشد."), 404
    storage.save("leaves", new)
    return jsonify(ok=True)


# =====================================================================
#  API — تنظیمات و داده‌ها
# =====================================================================
@app.route("/api/settings", methods=["GET", "POST"])
def api_settings():
    if request.method == "GET":
        return jsonify(ok=True, settings=storage.get_settings())
    d = request.get_json(force=True) or {}
    s = storage.get_settings()
    if "start_time" in d:
        s["start_time"] = str(d["start_time"]).strip() or s["start_time"]
    if "work_hours" in d:
        try:
            s["work_hours"] = max(1, min(24, float(d["work_hours"])))
        except Exception:
            pass
    if "threshold" in d:
        try:
            s["threshold"] = max(0.0, min(1.0, float(d["threshold"])))
        except Exception:
            pass
    if "company" in d:
        s["company"] = str(d["company"]).strip()
    storage.save("settings", s)
    if ENGINE is not None:
        ENGINE.cosine_threshold = float(s["threshold"])
    return jsonify(ok=True, settings=s)


@app.route("/api/backup")
def api_backup():
    """دانلود پشتیبان کامل به‌صورت یک فایل JSON."""
    payload = {
        "version": 1,
        "exported_at": datetime.now().isoformat(),
        "employees": storage.load("employees"),
        "attendance": storage.load("attendance"),
        "leaves": storage.load("leaves"),
        "approvals": storage.load("approvals"),
        "settings": storage.get_settings(),
    }
    import json as _json
    buf = io.BytesIO(_json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="attendance_backup.json", mimetype="application/json")


@app.route("/api/restore", methods=["POST"])
def api_restore():
    d = request.get_json(force=True) or {}
    for key in ("employees", "attendance", "leaves", "approvals"):
        if isinstance(d.get(key), list):
            storage.save(key, d[key])
    if isinstance(d.get("settings"), dict):
        s = storage.get_settings()
        s.update(d["settings"])
        storage.save("settings", s)
    return jsonify(ok=True)


@app.route("/api/reset", methods=["POST"])
def api_reset():
    storage.save("employees", [])
    storage.save("attendance", [])
    storage.save("leaves", [])
    storage.save("approvals", [])
    storage.save("settings", storage.DEFAULT_SETTINGS.copy())
    return jsonify(ok=True)


@app.route("/api/seed", methods=["POST"])
def api_seed():
    """ساخت دادهٔ نمونه برای آزمایش گزارش‌ها و داشبورد.

    توجه: کارمندان نمونه «چهرهٔ ثبت‌شده» ندارند؛ برای تشخیص در کیوسک باید
    از بخش «ثبت چهره» چهرهٔ واقعی هر فرد ثبت شود.
    """
    demo = [
        ("1001", "علی رضایی", "کارشناس فناوری اطلاعات", "فناوری اطلاعات"),
        ("1002", "سارا محمدی", "حسابدار ارشد", "مالی"),
        ("1003", "مهدی احمدی", "کارشناس فروش", "فروش"),
        ("1004", "فاطمه یوسفی", "مدیر منابع انسانی", "منابع انسانی"),
        ("1005", "حسین زمانی", "پشتیبان فنی", "فناوری اطلاعات"),
    ]
    emps = []
    for code, name, pos, unit in demo:
        emps.append({"id": storage.new_id(), "code": code, "name": name,
                     "position": pos, "unit": unit, "status": "فعال",
                     "encodings": [], "created": storage.now_ms()})
    storage.save("employees", emps)

    # تولید تردد تصادفی برای ۳۵ روز گذشته تا گزارش‌ها و داشبورد خالی نباشند.
    att = []
    today = datetime.now()
    for back in range(0, 35):
        day = today - timedelta(days=back)
        jd = storage.jalali_date_str(int(day.timestamp() * 1000))
        # جمعه تعطیل است و تردد ندارد؛ تعطیلی را از روی تقویم شمسی حساب می‌کنیم
        # (در jdatetime: شنبه=۰ ... جمعه=۶) نه از weekday میلادی.
        import jdatetime as _jd
        if _jd.date.fromgregorian(date=day.date()).weekday() == 6:
            continue
        for e in emps:
            if random.random() < 0.12:   # حدود ۱۲٪ مواقع، غیبت
                continue
            in_h, in_m = 8, random.choice([0, 5, 12, 20, 35, 50])
            if random.random() < 0.25:   # حدود ۲۵٪ مواقع، ورود با تأخیر (ساعت ۹)
                in_h, in_m = 9, random.choice([5, 15, 25])
            t_in = day.replace(hour=in_h, minute=in_m, second=0, microsecond=0)
            out_h = random.choice([16, 17, 17, 18])
            t_out = day.replace(hour=out_h, minute=random.choice([0, 10, 30]), second=0, microsecond=0)
            att.append({"id": storage.new_id(), "employee_id": e["id"], "type": "in",
                        "ts": int(t_in.timestamp() * 1000), "jdate": jd})
            att.append({"id": storage.new_id(), "employee_id": e["id"], "type": "out",
                        "ts": int(t_out.timestamp() * 1000), "jdate": jd})
    storage.save("attendance", att)
    storage.save("leaves", [])
    storage.save("approvals", [])
    return jsonify(ok=True, employees=len(emps), records=len(att))


if __name__ == "__main__":
    print("=" * 56)
    print("  سامانهٔ حضور و غیاب با تشخیص چهره")
    if ENGINE_ERROR:
        print("  هشدار:", ENGINE_ERROR)
    else:
        print("  موتور تشخیص چهره: آماده")
    print("  آدرس:  http://127.0.0.1:5000")
    print("=" * 56)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
